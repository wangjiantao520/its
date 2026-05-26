
import openpyxl
from pathlib import Path

excel_path = Path('/workspace/projects/assets') / '政企设备维保定额库及维保查勘问询信息记录表_20260526113332723.xlsx'

print('📊 开始分析Excel文件...\n')

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print('📋 Excel工作表列表:')
    print('------------------------')
    for idx, name in enumerate(wb.sheetnames, 1):
        print(f'{idx}. {name}')
    print('------------------------\n')

    # 详细分析每个工作表
    for sheet_name in wb.sheetnames:
        print(f'\n========================================')
        print(f'📄 工作表: {sheet_name}')
        print(f'========================================')
        
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f'📊 行数: {max_row}')
        print(f'📊 列数: {max_col}')
        
        # 显示前10行预览
        print('\n📋 数据预览 (前10行):')
        print('------------------------')
        for row_idx in range(1, min(11, max_row + 1)):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value) if cell.value is not None else '')
            print(f'行{row_idx}:', row_data)
        print('------------------------')
        
        # 如果是设备定额表，分析列名
        if '设备' in sheet_name or '定额' in sheet_name:
            print('\n🏷️  列名分析:')
            print('------------------------')
            if max_row >= 1:
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    print(f'列{col_idx}: {cell.value}')
            print('------------------------')
    
    print('\n✅ Excel分析完成！')
    
except Exception as e:
    print(f'❌ 分析Excel出错: {e}')
    import traceback
    traceback.print_exc()
